#!/usr/bin/env python3
"""
import_excel.py — ShipCII Dashboard
====================================
Import data dari 4 file Excel LANGSUNG ke Supabase.
Tidak perlu PostgreSQL lokal, tidak perlu migrate_to_supabase.py.

Urutan eksekusi:
  1. Update parameter kapal di tabel ship
  2. Import AIS tracking (Klasogun 82.625 baris, Balongan 6.511 baris)
  3. Import Noon Report Balongan (Juni 2026, 30 hari)
  4. Import Noon Report Klasogun (Juni 2025, 29 hari)
  5. Import data aggregat dari sheet FuelConsumptionByNOON REPORT
  6. Hitung dan simpan cii_daily (running CII harian)

Dependensi:
  pip install pandas openpyxl psycopg2-binary numpy
"""

import re
import math
import numpy as np
import pandas as pd
import psycopg2
from psycopg2.extras import execute_values
import openpyxl
from datetime import datetime, date, timedelta

# ─── CONFIG ──────────────────────────────────────────────────
DB_CONFIG = {
    "host":            "aws-1-ap-south-1.pooler.supabase.com",
    "port":            5432,
    "database":        "postgres",
    "user":            "postgres.qjqpepkgjfpbbwnvzuts",
    "password":        "TArahmat77!",
    "sslmode":         "require",
    "connect_timeout": 30,
}

EXCEL_BOTHSHIPS   = r"C:\Users\N1NRK\Downloads\TA_Rahmet\Data\AIS_2025_BothShips_FullYear_Updated (version 1)_YANG DIPAKE.xlsx"
EXCEL_NOON_BALONG = r"C:\Users\N1NRK\Downloads\TA_Rahmet\Data\NOON_REPORT_JUNE_2026_BALONGAN_FULL.xlsx"
EXCEL_NOON_KLAOS  = r"C:\Users\N1NRK\Downloads\TA_Rahmet\Data\NOON_REPORT_JUNE_2026_KLASOGUN_REVISED.xlsx"

MLR_B0 =  5.0801676028261635
MLR_B1 =  0.0030303477180683
MLR_B2 = -0.2783664235149623

CII_PARAMS = {
    "klasogun": {"dwt": 6627, "a": 5247, "c": 0.61, "Cf": 2.443},
    "balongan": {"dwt": 6736, "a": 5247, "c": 0.61, "Cf": 2.443},
}

REDUCTION_FACTOR = {
    2023: 0.05, 2024: 0.07, 2025: 0.09,
    2026: 0.11, 2027: 0.13625, 2028: 0.1625,
}

LAT_MIN, LAT_MAX = -11.0, -4.0
LON_MIN, LON_MAX = 103.5, 125.0

# ─── HELPERS ─────────────────────────────────────────────────

def get_ship_id(cur, ship_key):
    cur.execute("SELECT id FROM ship WHERE ship_key = %s", (ship_key,))
    row = cur.fetchone()
    if not row:
        raise ValueError(f"Ship '{ship_key}' tidak ditemukan.")
    return row[0]


def haversine_nm(lat1, lon1, lat2, lon2):
    R = 3440.065
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlam = math.radians(lon2 - lon1)
    a = math.sin(dphi/2)**2 + math.cos(phi1)*math.cos(phi2)*math.sin(dlam/2)**2
    return R * 2 * math.asin(math.sqrt(a))


def fix_timestamp_wib(ts_wib, date_val):
    if ts_wib is None or pd.isnull(ts_wib):
        return None
    try:
        ts = pd.to_datetime(ts_wib)
        if ts.year <= 1900:
            correct_date = pd.to_datetime(date_val).date()
            return datetime.combine(correct_date, ts.time())
        return ts.to_pydatetime()
    except Exception:
        return None


def calc_mlr_fuel(distance_nm, avg_speed):
    if distance_nm is None or avg_speed is None:
        return None
    result = MLR_B0 + MLR_B1 * float(distance_nm) + MLR_B2 * float(avg_speed)
    return round(max(0.0, result), 4)


def calc_cii_required(ship_key, year):
    p = CII_PARAMS[ship_key]
    reduction = REDUCTION_FACTOR.get(year, 0.09)
    cii_ref = p["a"] / (p["dwt"] ** p["c"])
    return round(cii_ref * (1 - reduction), 6)


def calc_grade(cii_attained, boundaries):
    if cii_attained is None:
        return "N/A"
    sup = boundaries.get("superior", 0)
    low = boundaries.get("lower", 0)
    upp = boundaries.get("upper", 0)
    inf = boundaries.get("inferior", 0)
    if   cii_attained < sup: return "A"
    elif cii_attained < low: return "B"
    elif cii_attained < upp: return "C"
    elif cii_attained < inf: return "D"
    else:                    return "E"


def get_boundaries(cur, ship_id, year):
    cur.execute("""
        SELECT boundary_superior, boundary_lower, boundary_upper, boundary_inferior
        FROM cii_boundaries
        WHERE ship_id = %s AND year = %s
        LIMIT 1
    """, (ship_id, year))
    row = cur.fetchone()
    if row:
        return {"superior": row[0], "lower": row[1], "upper": row[2], "inferior": row[3]}
    req = calc_cii_required(list(CII_PARAMS.keys())[0], year)
    return {
        "superior": req * 0.86,
        "lower":    req * 0.94,
        "upper":    req * 1.06,
        "inferior": req * 1.18,
    }


def parse_noon_sheet(ws):
    data = {}
    fuel_b35_values = []
    for row in ws.iter_rows(values_only=True):
        label = str(row[2]).strip() if row[2] is not None else ""
        col_d = row[3]
        val_e = row[4]
        val_h = row[7]
        if label == "VOYAGE NUMBER":
            data["voyage_number"] = str(val_e) if val_e else None
            if val_h and hasattr(val_h, "date"):
                data["voyage_date"] = val_h.date()
        elif label == "FROM":
            data["from_port"] = str(val_e).strip() if val_e else None
            data["to_port"]   = str(val_h).strip() if val_h else None
        elif label == "CURRENT LOCATION":
            data["current_location"] = str(val_e).strip() if val_e else None
        elif label == "DISTANCE TO RUN":
            try: data["distance_nm"] = float(val_e)
            except (TypeError, ValueError): pass
        elif label == "STEAMING TIME":
            try: data["steaming_time_h"] = float(val_e)
            except (TypeError, ValueError): pass
        elif label == "AVERAGE SPEED":
            try: data["avg_speed"] = float(val_e)
            except (TypeError, ValueError): pass
        elif label == "RPM (REV PER MINUTE)":
            try: data["rpm"] = float(val_e)
            except (TypeError, ValueError): pass
        elif label == "WEATHER CONDITION":
            data["weather"] = str(val_e).strip() if val_e else None
        if str(col_d).strip() == "B35/ME/AE" and val_e is not None:
            try:
                fuel_b35_values.append(float(val_e))
            except (TypeError, ValueError):
                pass
    if fuel_b35_values:
        data["fuel_cons_mt_per_day"] = fuel_b35_values[-1]
    if not data.get("voyage_date") and not data.get("distance_nm"):
        return None
    return data


# ─── STEP 1 ──────────────────────────────────────────────────

def update_ship_params(conn, cur):
    for ship_key, p in CII_PARAMS.items():
        year = 2025
        cii_ref = p["a"] / (p["dwt"] ** p["c"])
        cii_req = round(cii_ref * (1 - REDUCTION_FACTOR.get(year, 0.09)), 6)
        cur.execute("""
            UPDATE ship SET
                fuel_types      = %s,
                dwt             = %s,
                cii_param_a     = %s,
                cii_param_c     = %s,
                cii_ref_value   = %s,
                fuel_intercept  = %s,
                fuel_coef_speed = %s,
                fuel_coef_laden = %s
            WHERE ship_key = %s
        """, (
            ["B35", "B40", "B50"],
            p["dwt"], p["a"], p["c"],
            round(cii_req, 6),
            MLR_B0, MLR_B1, MLR_B2,
            ship_key,
        ))
    conn.commit()
    print("  ✅ Parameter kapal diupdate.")


# ─── STEP 2 ──────────────────────────────────────────────────

def import_ais(conn, cur):
    sheets = {
        "AISDATAMTKLASOGUN2025": "klasogun",
        "AISDATAMTBALONGAN2025": "balongan",
    }
    total = 0
    for sheet_name, ship_key in sheets.items():
        print(f"  Membaca {sheet_name}...")
        ship_id = get_ship_id(cur, ship_key)
        df = pd.read_excel(EXCEL_BOTHSHIPS, sheet_name=sheet_name)
        df.columns = [str(c).strip() for c in df.columns]
        rows_ok, rows_skip = [], 0
        for _, row in df.iterrows():
            try:
                lat = float(row["Latitude"])
                lon = float(row["Longitude"])
                if not (LAT_MIN <= lat <= LAT_MAX and LON_MIN <= lon <= LON_MAX):
                    rows_skip += 1; continue
                sog = float(row["SOG"]) if pd.notna(row.get("SOG")) else None
                if sog is not None and sog > 50:
                    rows_skip += 1; continue
                ts_utc = pd.to_datetime(row.get("TimeStamp"), errors="coerce", utc=True)
                if pd.isnull(ts_utc):
                    rows_skip += 1; continue
                ts_wib = fix_timestamp_wib(
                    row.get("TimeSTamp_WIB") or row.get("TimeStamp WIB"),
                    row.get("Date")
                )
                hdg = float(row["heading"]) if pd.notna(row.get("heading")) else None
                if hdg == 511: hdg = None
                cog = float(row["COG"]) if pd.notna(row.get("COG")) else None
                if cog == 360.0: cog = None
                rot_val = row.get("rot")
                rot = float(rot_val) if pd.notna(rot_val) else None
                if rot in (-128, -127, 127): rot = None
                navstatus = row.get("Navstatus")
                nav_int = int(navstatus) if pd.notna(navstatus) else None
                mmsi_val = row.get("MMSI")
                mmsi = int(mmsi_val) if pd.notna(mmsi_val) else None
                ais_class = str(row.get("class", "A")).strip()
                rows_ok.append((
                    ship_id, None,
                    str(mmsi) if mmsi else None,
                    ts_utc, ts_wib,
                    lat, lon, sog, cog, hdg,
                    str(nav_int) if nav_int is not None else None,
                    None, None, False,
                    mmsi, ais_class,
                ))
            except (TypeError, ValueError, KeyError):
                rows_skip += 1
                continue
        if not rows_ok:
            print(f"    Tidak ada data valid untuk {ship_key}.")
            continue
        sql = """
            INSERT INTO ais_tracking
                (ship_id, voyage_id, ais_record_id, base_datetime,
                 timestamp_wib, lat, lon, sog, cog, heading,
                 status, cargo_status, ais_point_type, is_estimated,
                 mmsi, ais_class)
            VALUES %s
            ON CONFLICT DO NOTHING
        """
        execute_values(cur, sql, rows_ok, page_size=2000)
        conn.commit()
        print(f"    ✅ {len(rows_ok):,} baris AIS {ship_key} ({rows_skip:,} dibuang).")
        total += len(rows_ok)
    return total


# ─── STEP 3 & 4 ──────────────────────────────────────────────

def import_noon_report(conn, cur, excel_path, ship_key):
    ship_id = get_ship_id(cur, ship_key)
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    rows_ok, skipped = [], []
    for sheet_name in wb.sheetnames:
        clean = sheet_name.strip()
        if not re.match(r"^\d{2}\.\d{2}\.\d{4}$", clean):
            skipped.append(sheet_name); continue
        ws  = wb[sheet_name]
        rec = parse_noon_sheet(ws)
        if not rec:
            skipped.append(sheet_name); continue
        if not rec.get("voyage_date"):
            try:
                d, m, y = clean.split(".")
                rec["voyage_date"] = date(int(y), int(m), int(d))
            except ValueError:
                skipped.append(sheet_name); continue
        dist  = rec.get("distance_nm")
        speed = rec.get("avg_speed")
        rows_ok.append((
            ship_id,
            rec["voyage_date"],
            rec.get("from_port"),
            rec.get("to_port"),
            speed, dist,
            rec.get("steaming_time_h"),
            None,
            rec.get("fuel_cons_mt_per_day"),
            "B-35",
            calc_mlr_fuel(dist, speed),
            rec.get("rpm"),
            rec.get("weather"),
            rec.get("current_location"),
        ))
    wb.close()
    if not rows_ok:
        print(f"    Tidak ada sheet valid.")
        return 0
    sql = """
        INSERT INTO noon_report
            (ship_id, voyage_date, from_port, to_port,
             avg_speed, distance_nm, steaming_time_h, cargo_status,
             fuel_cons_mt_per_day, fuel_type, fuel_cons_mlr,
             rpm, weather, current_location)
        VALUES %s
        ON CONFLICT (ship_id, voyage_date) DO NOTHING
    """
    execute_values(cur, sql, rows_ok, page_size=500)
    conn.commit()
    print(f"    ✅ {len(rows_ok)} hari noon report {ship_key} ({len(skipped)} sheet dilewati).")
    return len(rows_ok)


# ─── STEP 5 ──────────────────────────────────────────────────

def import_fuel_aggregat(conn, cur):
    df = pd.read_excel(EXCEL_BOTHSHIPS, sheet_name="FuelConsumptionByNOON REPORT", header=None)
    total = 0
    for ship_key, col_date, col_dist, col_speed, col_fuel in [
        ("balongan", 1, 2, 3, 4),
        ("klasogun", 6, 7, 8, 9),
    ]:
        ship_id = get_ship_id(cur, ship_key)
        rows_ok = []
        for _, row in df.iterrows():
            try:
                voyage_date = pd.to_datetime(row[col_date])
                if pd.isnull(voyage_date): continue
                dist  = float(row[col_dist])
                speed = float(str(row[col_speed]).replace(",", "."))
                fuel  = float(row[col_fuel])
                rows_ok.append((
                    ship_id, voyage_date.date(),
                    None, None,
                    speed, dist, None, None,
                    fuel, "B-35", calc_mlr_fuel(dist, speed),
                    None, None, None,
                ))
            except (TypeError, ValueError):
                continue
        if not rows_ok: continue
        sql = """
            INSERT INTO noon_report
                (ship_id, voyage_date, from_port, to_port,
                 avg_speed, distance_nm, steaming_time_h, cargo_status,
                 fuel_cons_mt_per_day, fuel_type, fuel_cons_mlr,
                 rpm, weather, current_location)
            VALUES %s
            ON CONFLICT (ship_id, voyage_date) DO NOTHING
        """
        execute_values(cur, sql, rows_ok, page_size=500)
        conn.commit()
        total += len(rows_ok)
        print(f"    ✅ {len(rows_ok)} baris aggregat {ship_key}.")
    return total


# ─── STEP 6 ──────────────────────────────────────────────────

def calc_daily_distance(cur, ship_id, target_date):
    cur.execute("""
        SELECT lat, lon FROM ais_tracking
        WHERE ship_id = %s
          AND (base_datetime AT TIME ZONE 'UTC')::date = %s
          AND sog > 0.5
        ORDER BY base_datetime
    """, (ship_id, target_date))
    points = cur.fetchall()
    if len(points) < 2:
        return 0.0
    total_dist = 0.0
    for i in range(1, len(points)):
        lat1, lon1 = float(points[i-1][0]), float(points[i-1][1])
        lat2, lon2 = float(points[i][0]),   float(points[i][1])
        d = haversine_nm(lat1, lon1, lat2, lon2)
        if d < 100:
            total_dist += d
    return round(total_dist, 4)


def build_cii_daily(conn, cur):
    for ship_key, params in CII_PARAMS.items():
        ship_id = get_ship_id(cur, ship_key)
        dwt = params["dwt"]
        Cf  = params["Cf"]
        print(f"  Menghitung cii_daily {ship_key}...")

        cur.execute("""
            SELECT
              MIN((base_datetime AT TIME ZONE 'UTC')::date),
              MAX((base_datetime AT TIME ZONE 'UTC')::date)
            FROM ais_tracking WHERE ship_id = %s
        """, (ship_id,))
        row = cur.fetchone()
        if not row or not row[0]:
            print(f"    Tidak ada data AIS untuk {ship_key}."); continue

        start_date, end_date = row[0], row[1]
        year = start_date.year
        cii_required = calc_cii_required(ship_key, year)
        boundaries   = get_boundaries(cur, ship_id, year)

        cur.execute("""
            SELECT voyage_date, distance_nm, fuel_cons_mt_per_day, fuel_cons_mlr
            FROM noon_report
            WHERE ship_id = %s AND voyage_date BETWEEN %s AND %s
            ORDER BY voyage_date
        """, (ship_id, start_date, end_date))
        noon_map = {r[0]: r for r in cur.fetchall()}

        dist_ytd = fuel_ytd = co2_ytd = tw_ytd = 0.0
        rows_daily = []
        current = start_date

        while current <= end_date:
            dist_day = calc_daily_distance(cur, ship_id, current)

            fuel_day = None
            if current in noon_map:
                nr = noon_map[current]
                fuel_day = float(nr[2]) if nr[2] is not None else (
                           float(nr[3]) if nr[3] is not None else None)
            if fuel_day is None and dist_day > 0:
                cur.execute("""
                    SELECT AVG(sog) FROM ais_tracking
                    WHERE ship_id = %s
                      AND (base_datetime AT TIME ZONE 'UTC')::date = %s
                      AND sog > 0.5
                """, (ship_id, current))
                avg_sog = cur.fetchone()[0]
                if avg_sog:
                    fuel_day = calc_mlr_fuel(dist_day, float(avg_sog))

            fuel_day = fuel_day or 0.0
            co2_day  = fuel_day * Cf * 1_000_000
            tw_day   = dwt * dist_day

            dist_ytd += dist_day
            fuel_ytd += fuel_day
            co2_ytd  += co2_day
            tw_ytd   += tw_day

            running_cii = (co2_ytd / tw_ytd) if tw_ytd > 0 else None
            grade = calc_grade(running_cii, boundaries) if running_cii else None

            day_of_year  = (current - date(year, 1, 1)).days + 1
            days_in_year = 366 if year % 4 == 0 else 365
            projected = date_limit = days_to_lim = None

            if running_cii and day_of_year > 30:
                cii_per_day = running_cii / day_of_year
                remaining   = days_in_year - day_of_year
                projected   = round(running_cii + cii_per_day * remaining, 6)
                if cii_per_day > 0 and running_cii < cii_required:
                    days_needed = (cii_required - running_cii) / cii_per_day
                    if days_needed < 365:
                        date_limit  = current + timedelta(days=int(days_needed))
                        days_to_lim = int(days_needed)

            rows_daily.append((
                ship_id, current, year, current.month,
                round(dist_day, 4),
                round(fuel_day, 6),
                round(co2_day, 2),
                round(dist_ytd, 4),
                round(fuel_ytd, 6),
                round(co2_ytd, 2),
                round(tw_ytd, 2),
                round(running_cii, 6) if running_cii else None,
                round(cii_required, 6),
                grade,
                round(projected, 6) if projected else None,
                days_to_lim,
                date_limit,
            ))
            current += timedelta(days=1)

        if rows_daily:
            # PERBAIKAN: ON CONFLICT pakai (ship_id, date) — sesuai UNIQUE constraint di tabel
            sql = """
                INSERT INTO cii_daily (
                    ship_id, date, year, month,
                    distance_nm_day, fuel_cons_mt_day, co2_emission_g_day,
                    distance_nm_ytd, fuel_cons_mt_ytd, co2_emission_g_ytd,
                    transport_work_ytd,
                    running_cii, cii_required, running_grade,
                    projected_cii_eoy, days_to_limit, date_limit_reached
                ) VALUES %s
                ON CONFLICT (ship_id, date) DO NOTHING
            """
            execute_values(cur, sql, rows_daily, page_size=500)
            conn.commit()
            print(f"    ✅ {len(rows_daily)} hari cii_daily {ship_key} tersimpan.")


# ─── MAIN ─────────────────────────────────────────────────────

def main():
    conn = psycopg2.connect(**DB_CONFIG)
    cur  = conn.cursor()

    print("\n[1/6] Update parameter kapal...")
    update_ship_params(conn, cur)

    print("\n[2/6] Import AIS tracking...")
    n_ais = import_ais(conn, cur)
    print(f"  Total: {n_ais:,} baris AIS")

    print("\n[3/6] Import Noon Report Balongan (Juni 2026)...")
    n_nr1 = import_noon_report(conn, cur, EXCEL_NOON_BALONG, "balongan")

    print("\n[4/6] Import Noon Report Klasogun (Juni 2025)...")
    n_nr2 = import_noon_report(conn, cur, EXCEL_NOON_KLAOS, "klasogun")

    print("\n[5/6] Import aggregat fuel dari FuelConsumptionByNOON REPORT...")
    n_agg = import_fuel_aggregat(conn, cur)

    print("\n[6/6] Hitung dan simpan cii_daily (running CII harian)...")
    print("  (Proses ini memakan waktu 5-15 menit karena kalkulasi Haversine per hari)")
    build_cii_daily(conn, cur)

    cur.close()
    conn.close()

    print(f"\n{'='*55}")
    print(f"✅ Import selesai.")
    print(f"   AIS tracking : {n_ais:,} baris")
    print(f"   Noon report  : {n_nr1 + n_nr2 + n_agg} hari")
    print(f"{'='*55}")
    print("\nVerifikasi di Supabase SQL Editor:")
    print("""
  SELECT 'ship'        AS tbl, COUNT(*) FROM ship        UNION ALL
  SELECT 'noon_report',        COUNT(*) FROM noon_report  UNION ALL
  SELECT 'cii_daily',          COUNT(*) FROM cii_daily    UNION ALL
  SELECT 'ais_tracking',       COUNT(*) FROM ais_tracking;
    """)


if __name__ == "__main__":
    main()